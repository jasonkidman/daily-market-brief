"""Local S&P 500 constituent reference validation and safe replacement."""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
import os
import re
from tempfile import NamedTemporaryFile
from typing import Callable, Iterable


REQUIRED_FIELDS = ("source_ticker", "yahoo_ticker", "name", "as_of_date")
NON_EQUITY_WORDS = ("cash", "currency", "future", "futures", "option", "swap")


class ConstituentReferenceError(ValueError):
    """Raised when a constituent reference is unsafe to use or replace."""


def normalize_yahoo_ticker(source_ticker: str) -> str:
    """Apply the documented Yahoo class-share punctuation conversion."""
    return str(source_ticker or "").strip().upper().replace(".", "-").replace("/", "-")


def _as_of_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError as exc:
        raise ConstituentReferenceError("as_of_date 无效。") from exc


def _normalised_row(row: dict) -> dict:
    source_ticker = str(row.get("source_ticker", "")).strip().upper()
    yahoo_ticker = normalize_yahoo_ticker(row.get("yahoo_ticker") or source_ticker)
    return {
        "source_ticker": source_ticker,
        "yahoo_ticker": yahoo_ticker,
        "name": str(row.get("name", "")).strip(),
        "as_of_date": _as_of_date(row.get("as_of_date")),
    }


def validate_constituents(rows: Iterable[dict], minimum_count: int = 480) -> list[dict]:
    """Validate equity-only rows before a new reference can replace the old one."""
    cleaned, seen = [], set()
    for raw in rows:
        row = _normalised_row(raw)
        if not row["source_ticker"] or not row["yahoo_ticker"] or not row["name"]:
            raise ConstituentReferenceError("成分股 ticker 或名称不能为空。")
        name = row["name"].casefold()
        if any(word in name for word in NON_EQUITY_WORDS):
            continue
        if row["yahoo_ticker"] in seen:
            raise ConstituentReferenceError("Yahoo ticker 重复。")
        seen.add(row["yahoo_ticker"])
        cleaned.append(row)
    if len(cleaned) < minimum_count:
        raise ConstituentReferenceError(f"股票证券数量不足：{len(cleaned)}，至少需要 {minimum_count}。")
    return cleaned


def load_constituents(path: Path) -> list[dict]:
    """Read the previously validated local reference file used by daily generation."""
    path = Path(path)
    try:
        with path.open(encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames or not set(REQUIRED_FIELDS).issubset(reader.fieldnames):
                raise ConstituentReferenceError("成分股 reference CSV 缺少必要字段。")
            return [_normalised_row(row) for row in reader]
    except FileNotFoundError as exc:
        raise ConstituentReferenceError("成分股 reference CSV 不存在。") from exc


def update_reference(reference_path: Path, fetcher: Callable[[], Iterable[dict]],
                     minimum_count: int = 480) -> bool:
    """Fetch, validate and atomically replace a local reference only on success."""
    reference_path = Path(reference_path)
    cleaned = validate_constituents(list(fetcher()), minimum_count=minimum_count)
    reference_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", newline="", dir=reference_path.parent,
                            prefix=f".{reference_path.name}.", suffix=".tmp", delete=False) as handle:
        temporary_path = Path(handle.name)
        writer = csv.DictWriter(handle, fieldnames=REQUIRED_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(cleaned)
    try:
        os.replace(temporary_path, reference_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return True


def parse_state_street_holdings(workbook_bytes: bytes) -> list[dict]:
    """Parse State Street's official daily SPY holdings workbook into raw rows."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is installed in normal runs
        raise ConstituentReferenceError("需要 openpyxl 解析 State Street holdings 文件。") from exc

    from io import BytesIO

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    header_index = next((index for index, row in enumerate(values[:60])
                         if any(str(cell or "").strip().casefold() == "ticker" for cell in row)), None)
    if header_index is None:
        raise ConstituentReferenceError("State Street holdings 文件未找到 Ticker 列。")
    header = [str(cell or "").strip().casefold() for cell in values[header_index]]
    ticker_index = header.index("ticker")
    name_index = next((index for index, value in enumerate(header)
                       if value in {"name", "security name", "security"}), None)
    if name_index is None:
        raise ConstituentReferenceError("State Street holdings 文件未找到 Name 列。")
    as_of = _state_street_as_of_date(values[:header_index + 1])
    parsed = []
    for row in values[header_index + 1:]:
        ticker = str(row[ticker_index] or "").strip()
        name = str(row[name_index] or "").strip()
        if ticker and name:
            parsed.append({"source_ticker": ticker, "name": name, "as_of_date": as_of})
    return parsed


def _state_street_as_of_date(rows: Iterable[tuple]) -> str:
    for row in rows:
        for cell in row:
            if isinstance(cell, (date, datetime)):
                return _as_of_date(cell)
            text = str(cell or "")
            match = re.search(
                r"(?:as\s+of\s*)?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2}|\d{1,2}-[A-Za-z]{3}-\d{4})",
                text,
                re.I,
            )
            if match:
                candidate = match.group(1).replace("/", "-")
                for pattern in ("%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d", "%d-%b-%Y"):
                    try:
                        return datetime.strptime(candidate, pattern).date().isoformat()
                    except ValueError:
                        continue
    raise ConstituentReferenceError("State Street holdings 文件未找到有效 as_of_date。")
